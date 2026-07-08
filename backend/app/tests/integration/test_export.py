import io
import zipfile
from datetime import UTC, datetime, timedelta

from openpyxl import load_workbook

from app.tests.integration.conftest import bearer, login

TODAY = datetime.now(UTC).date()


async def seed(client, headers):
    habit = (
        await client.post(
            "/habits", json={"name": "Meditate", "question": "Did you meditate?"}, headers=headers
        )
    ).json()
    numeric = (
        await client.post(
            "/habits", json={"name": "Read", "type": 1, "target_value": 20, "unit": "pages"},
            headers=headers,
        )
    ).json()
    for offset in range(3):
        d = TODAY - timedelta(days=offset)
        await client.post(f"/habits/{habit['id']}/entries/{d}/toggle", headers=headers)
    await client.put(
        f"/habits/{numeric['id']}/entries/{TODAY}",
        json={"value": 25000, "notes": "long session"},
        headers=headers,
    )
    return habit, numeric


async def test_csv_export_layout(client):
    headers = bearer(await login(client, 4000))
    await seed(client, headers)

    response = await client.get("/export/csv", headers=headers)
    assert response.status_code == 200
    zf = zipfile.ZipFile(io.BytesIO(response.content))
    names = zf.namelist()
    assert "Habits.csv" in names
    assert "001 Meditate/Checkmarks.csv" in names
    assert "001 Meditate/Scores.csv" in names
    assert "002 Read/Checkmarks.csv" in names

    habits_csv = zf.read("Habits.csv").decode()
    assert "Meditate" in habits_csv and "NUMERICAL" in habits_csv

    checkmarks = zf.read("001 Meditate/Checkmarks.csv").decode()
    assert "YES_MANUAL" in checkmarks


async def test_csv_round_trip(client):
    headers = bearer(await login(client, 4001))
    await seed(client, headers)
    exported = (await client.get("/export/csv", headers=headers)).content

    # Import into a fresh user.
    headers_b = bearer(await login(client, 4002))
    response = await client.post(
        "/import/csv",
        files={"file": ("export.zip", exported, "application/zip")},
        headers=headers_b,
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["habits_created"] == 2
    assert body["entries_imported"] == 4  # 3 toggles + 1 numeric

    habits = (await client.get("/habits", headers=headers_b)).json()
    assert {h["name"] for h in habits} == {"Meditate", "Read"}

    numeric = next(h for h in habits if h["name"] == "Read")
    assert numeric["type"] == 1
    assert numeric["target_value"] == 20.0

    entries = await client.get(
        f"/habits/{numeric['id']}/entries",
        params={"from": str(TODAY), "to": str(TODAY)},
        headers=headers_b,
    )
    assert entries.json()[0]["value"] == 25000
    assert entries.json()[0]["notes"] == "long session"


async def test_import_is_idempotent_by_name(client):
    headers = bearer(await login(client, 4003))
    await seed(client, headers)
    exported = (await client.get("/export/csv", headers=headers)).content

    first = (
        await client.post(
            "/import/csv", files={"file": ("e.zip", exported, "application/zip")}, headers=headers
        )
    ).json()
    assert first["habits_created"] == 0  # names already exist

    habits = (await client.get("/habits", headers=headers)).json()
    assert len(habits) == 2


def _loop_zip() -> bytes:
    d0, d1, d2, d3 = (TODAY - timedelta(days=n) for n in range(4))
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w") as zf:
        zf.writestr(
            "Habits.csv",
            "Position,Name,Question,Description,NumRepetitions,Interval,Color\n"
            "001,sadaqa,,,1,1,#F57C00\n"
            "002,cvn,,,1,7,#303030\n",
        )
        # Real Loop per-habit files have no header row and are newest-first.
        zf.writestr("001 sadaqa/Checkmarks.csv", f"{d0},2\n{d1},1\n{d2},0\n{d3},2\n")
        zf.writestr("001 sadaqa/Scores.csv", f"{d0},1.0000\n")
        # Root aggregate files must be ignored.
        zf.writestr("Checkmarks.csv", f"Date,sadaqa,\n{d0},2,\n")
        zf.writestr("Scores.csv", f"Date,sadaqa,\n{d0},1.0000,\n")
    return buffer.getvalue()


async def test_import_loop_android_format(client):
    headers = bearer(await login(client, 4007))
    response = await client.post(
        "/import/csv",
        files={"file": ("loop.zip", _loop_zip(), "application/zip")},
        headers=headers,
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["habits_created"] == 2
    assert body["entries_imported"] == 2  # only the two value-2 rows; 1 and 0 skipped

    habits = (await client.get("/habits", headers=headers)).json()
    by_name = {h["name"]: h for h in habits}
    assert set(by_name) == {"sadaqa", "cvn"}
    assert by_name["sadaqa"]["color"] == 2  # #F57C00
    assert by_name["cvn"]["freq_num"] == 1
    assert by_name["cvn"]["freq_den"] == 7
    assert by_name["cvn"]["color"] == 8  # non-palette hex falls back to default

    entries = await client.get(
        f"/habits/{by_name['sadaqa']['id']}/entries",
        params={"from": str(TODAY), "to": str(TODAY)},
        headers=headers,
    )
    assert entries.json()[0]["value"] == 2  # newest row imported despite no header


async def test_export_telegram_requires_bot_link(client):
    headers = bearer(await login(client, 4100))
    response = await client.get("/export/xlsx", params={"to_telegram": "1"}, headers=headers)
    assert response.status_code == 422


async def test_export_telegram_delivers(client, monkeypatch):
    from sqlalchemy import update

    from app.api.routers import export as export_router
    from app.infrastructure.db.tables import UserRow

    headers = bearer(await login(client, 4101))
    async with client.session_factory() as session:
        await session.execute(
            update(UserRow).where(UserRow.telegram_id == 4101).values(bot_linked=True)
        )
        await session.commit()

    sent = {}

    async def fake_send(bot_token, chat_id, filename, content, caption=None):
        sent.update(chat_id=chat_id, filename=filename, size=len(content))
        return True

    monkeypatch.setattr(export_router, "send_document", fake_send)

    response = await client.get("/export/xlsx", params={"to_telegram": "1"}, headers=headers)
    assert response.status_code == 200
    assert response.json() == {"ok": True, "delivered": "telegram"}
    assert sent["chat_id"] == 4101
    assert sent["filename"] == "habitflow-export.xlsx"
    assert sent["size"] > 0


async def test_import_rejects_garbage(client):
    headers = bearer(await login(client, 4004))
    response = await client.post(
        "/import/csv", files={"file": ("junk.zip", b"not a zip", "application/zip")}, headers=headers
    )
    assert response.status_code == 422


async def test_xlsx_export(client):
    headers = bearer(await login(client, 4005))
    await seed(client, headers)

    response = await client.get("/export/xlsx", headers=headers)
    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.content))
    assert "Habits" in workbook.sheetnames
    assert "Meditate" in workbook.sheetnames
    habits_sheet = workbook["Habits"]
    names = [row[1] for row in habits_sheet.iter_rows(min_row=2, values_only=True)]
    assert names == ["Meditate", "Read"]


async def test_xlsx_analytics_sheets(client):
    headers = bearer(await login(client, 4006))
    await seed(client, headers)

    response = await client.get("/export/xlsx", headers=headers)
    workbook = load_workbook(io.BytesIO(response.content))
    for name in ("Summary", "Weekly", "Monthly"):
        assert name in workbook.sheetnames

    summary = workbook["Summary"]
    header = [c.value for c in summary[1]]
    assert header == [
        "Habit", "Type", "Success rate", "Current streak", "Best streak",
        "Total completions", "Total value", "Daily average", "First entry", "Last entry",
    ]
    rows = {row[0]: row for row in summary.iter_rows(min_row=2, values_only=True)}
    # 3 daily toggles → current streak 3, 3 completions.
    assert rows["Meditate"][3] == 3
    assert rows["Meditate"][5] == 3
    # 25000 millis = 25.0 total, one entry day → daily average 25.0, meets target 20.
    assert rows["Read"][5] == 1
    assert rows["Read"][6] == 25.0
    assert rows["Read"][7] == 25.0

    weekly = workbook["Weekly"]
    assert [c.value for c in weekly[1]] == [
        "Habit", "Week", "Completions", "Success rate", "Total value", "Daily average", "Avg score",
    ]
    weekly_rows = list(weekly.iter_rows(min_row=2, values_only=True))
    assert sum(r[2] for r in weekly_rows if r[0] == "Meditate") == 3
    assert sum(r[4] for r in weekly_rows if r[0] == "Read") == 25.0
    # Entries cover every expected day, so success rates are in (0, 1].
    # Exact values depend on where TODAY falls in the ISO week, so assert bounds only.
    for r in weekly_rows:
        assert 0 < r[3] <= 1.0
    read_row = next(r for r in weekly_rows if r[0] == "Read")
    assert read_row[5] == 25.0  # daily average

    monthly = workbook["Monthly"]
    monthly_rows = list(monthly.iter_rows(min_row=2, values_only=True))
    assert sum(r[2] for r in monthly_rows if r[0] == "Meditate") == 3
