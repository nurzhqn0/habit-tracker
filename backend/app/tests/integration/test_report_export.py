import io
from datetime import UTC, datetime, timedelta

from openpyxl import load_workbook

from app.tests.integration.conftest import bearer, login

TODAY = datetime.now(UTC).date()


def workbook_from(response):
    assert response.status_code == 200, response.text
    return load_workbook(io.BytesIO(response.content))


def sheet_rows(ws):
    return [list(row) for row in ws.iter_rows(values_only=True)]


async def make_room(client, headers, name="Report Crew"):
    response = await client.post("/rooms", json={"name": name}, headers=headers)
    assert response.status_code == 201, response.text
    return response.json()


async def seed_personal(client, headers):
    habit = (
        await client.post(
            "/habits", json={"name": "Meditate", "question": "Did you meditate?"}, headers=headers
        )
    ).json()
    numeric = (
        await client.post(
            "/habits", json={"name": "Read", "type": 1, "target_value": 20, "unit": "pages"},
            headers=headers,
        )
    ).json()
    for offset in range(3):
        d = TODAY - timedelta(days=offset)
        await client.post(f"/habits/{habit['id']}/entries/{d}/toggle", headers=headers)
    await client.put(f"/habits/{numeric['id']}/entries/{TODAY}", json={"value": 25000}, headers=headers)
    return habit, numeric


async def test_personal_report_defaults_to_week(client):
    headers = bearer(await login(client, 6000))
    await seed_personal(client, headers)

    wb = workbook_from(await client.get("/export/report/xlsx", headers=headers))
    assert wb.sheetnames == ["Daily"]
    rows = sheet_rows(wb["Daily"])

    expected_dates = [(TODAY - timedelta(days=6 - i)).isoformat() for i in range(7)]
    assert rows[0] == ["Habit", *expected_dates, "Total"]

    meditate = next(r for r in rows if r[0] == "Meditate")
    assert meditate[-4:] == ["✓", "✓", "✓", 3]  # last 3 days completed

    read = next(r for r in rows if r[0] == "Read")
    assert read[-2:] == [25, 25.0]  # today's value and numeric period total

    totals = rows[-1]
    assert totals[0] == "Totals"
    assert totals[-1] == 4  # 3 Meditate completions + 1 Read completion


async def test_personal_report_range_and_validation(client):
    headers = bearer(await login(client, 6001))
    habit, _ = await seed_personal(client, headers)
    await client.delete(f"/habits/{habit['id']}", headers=headers)

    response = await client.get(
        "/export/report/xlsx",
        params={"from": str(TODAY - timedelta(days=2)), "to": str(TODAY)},
        headers=headers,
    )
    rows = sheet_rows(workbook_from(response)["Daily"])
    assert len(rows[0]) == 5  # Habit + 3 dates + Total
    assert all(r[0] != "Meditate" for r in rows)  # deleted habit excluded

    bad = await client.get(
        "/export/report/xlsx",
        params={"from": str(TODAY), "to": str(TODAY - timedelta(days=1))},
        headers=headers,
    )
    assert bad.status_code == 422

    too_long = await client.get(
        "/export/report/xlsx",
        params={"from": str(TODAY - timedelta(days=400)), "to": str(TODAY)},
        headers=headers,
    )
    assert too_long.status_code == 422


async def backdate_joins(client, room_id, days=30):
    """Room stats clamp history to joined_at; pretend everyone joined earlier."""
    from sqlalchemy import update

    from app.infrastructure.db.tables import RoomMemberRow

    joined = datetime.now(UTC).replace(tzinfo=None) - timedelta(days=days)
    async with client.session_factory() as session:
        await session.execute(
            update(RoomMemberRow).where(RoomMemberRow.room_id == room_id).values(joined_at=joined)
        )
        await session.commit()


async def seed_room(client, owner, member):
    room = await make_room(client, owner)
    await client.post("/rooms/join", json={"code": room["invite_code"]}, headers=member)
    await backdate_joins(client, room["id"])
    room_habit = (
        await client.post(f"/rooms/{room['id']}/habits", json={"name": "Pushups"}, headers=owner)
    ).json()
    linked = (
        await client.post(f"/rooms/{room['id']}/habits/{room_habit['id']}/link", json={}, headers=member)
    ).json()["habit_id"]
    await client.post(f"/habits/{linked}/entries/{TODAY}/toggle", headers=member)
    await client.post(f"/habits/{linked}/entries/{TODAY - timedelta(days=1)}/toggle", headers=member)
    return room, linked


async def test_room_report_requires_admin(client):
    owner = bearer(await login(client, 6010))
    member = bearer(await login(client, 6011))
    outsider = bearer(await login(client, 6012))
    room, _ = await seed_room(client, owner, member)

    assert (await client.get(f"/rooms/{room['id']}/export/xlsx", headers=member)).status_code == 403
    assert (await client.get(f"/rooms/{room['id']}/export/xlsx", headers=outsider)).status_code == 404
    assert (await client.get(f"/rooms/{room['id']}/export/xlsx", headers=owner)).status_code == 200


async def test_room_report_sheets(client):
    owner = bearer(await login(client, 6020))
    member = bearer(await login(client, 6021))
    room, _ = await seed_room(client, owner, member)
    idle = bearer(await login(client, 6022))  # joins but links nothing
    await client.post("/rooms/join", json={"code": room["invite_code"]}, headers=idle)

    wb = workbook_from(await client.get(f"/rooms/{room['id']}/export/xlsx", headers=owner))
    assert wb.sheetnames == ["Daily", "Analytics", "Leaderboard"]

    daily = sheet_rows(wb["Daily"])
    assert daily[0][:2] == ["Member", "Habit"]
    member_row = next(r for r in daily if r[0] == "Alice (@alice6021)")
    assert member_row[1] == "Pushups"
    assert member_row.count("✓") == 2
    assert member_row[-1] == 2
    assert all(r[0] != "Alice (@alice6022)" for r in daily)  # no link, no daily row

    analytics = sheet_rows(wb["Analytics"])
    assert analytics[0] == [
        "Member", "Habit", "Target", "Score", "Current streak", "Completions", "Success rate",
    ]
    row = next(r for r in analytics if r[0] == "Alice (@alice6021)")
    assert row[1] == "Pushups"
    assert row[2] == "every day"
    assert row[3] > 0
    assert row[4] == 2
    assert row[5] == 2
    assert 0 < row[6] <= 1

    board = sheet_rows(wb["Leaderboard"])
    assert len(board) == 4  # header + all 3 members, linked or not
    assert board[1][:2] == [1, "Alice"]
    assert board[1][5] == 2  # member's completions lead
    idle_rows = [r for r in board[1:] if r[2] == "alice6022"]
    assert idle_rows and idle_rows[0][6] == 0  # idle member listed with zero linked habits


async def test_room_report_window_limits_completions(client):
    owner = bearer(await login(client, 6030))
    member = bearer(await login(client, 6031))
    room, linked = await seed_room(client, owner, member)
    await client.post(
        f"/habits/{linked}/entries/{TODAY - timedelta(days=10)}/toggle", headers=member
    )

    def completions(wb):
        analytics = sheet_rows(wb["Analytics"])
        board = sheet_rows(wb["Leaderboard"])
        return analytics[1][5], board[1][5]

    wb = workbook_from(await client.get(f"/rooms/{room['id']}/export/xlsx", headers=owner))
    assert completions(wb) == (2, 2)  # default week window excludes the old entry

    wb = workbook_from(
        await client.get(
            f"/rooms/{room['id']}/export/xlsx",
            params={"from": str(TODAY), "to": str(TODAY)},
            headers=owner,
        )
    )
    assert completions(wb) == (1, 1)
