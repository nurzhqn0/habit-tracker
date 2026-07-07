"""Period-scoped XLSX reports: personal daily grid and room daily/analytics/leaderboard.

Room sheets count completions with rooms._is_success (leaderboard parity, SKIP counts);
the personal sheet uses csv_io._is_completed (backup-export parity, SKIP excluded).
"""

import io
from datetime import date as Date
from datetime import timedelta

from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.application import habit_math
from app.application.use_cases.csv_io import _is_completed, _style_headers
from app.application.use_cases.rooms import _is_success, _require_admin, leaderboard
from app.domain.errors import NotFoundError, ValidationError
from app.domain.models.entry import NO, SKIP, YES_AUTO, YES_MANUAL
from app.domain.services.entry_list import get_known
from app.infrastructure.repositories.habit_repo import EntryRepo, HabitRepo
from app.infrastructure.repositories.room_repo import RoomRepo
from app.infrastructure.repositories.user_repo import UserRepo

MAX_REPORT_DAYS = 366  # one column per day; cap keeps the Daily sheet usable


def _resolve_window(from_date: Date | None, to_date: Date | None, today: Date) -> tuple[Date, Date]:
    to = to_date or today
    frm = from_date or to - timedelta(days=6)
    if frm > to:
        raise ValidationError("'from' must be on or before 'to'")
    if (to - frm).days + 1 > MAX_REPORT_DAYS:
        raise ValidationError(f"Date range too large (max {MAX_REPORT_DAYS} days)")
    return frm, to


def _day_cell(habit_row, value: int | None) -> str | float | None:
    if value is None:
        return None
    if habit_row.type == 1:
        return value / 1000 if value >= 0 else None
    if value in (YES_MANUAL, YES_AUTO):
        return "✓"
    if value == SKIP:
        return "—"
    if value == NO:
        return "✗"
    return None


def _format_target(habit_row) -> str:
    if habit_row.type == 1:
        kind = "at most" if habit_row.target_type == 1 else "at least"
        return f"{kind} {habit_row.target_value:g} {habit_row.unit}".strip()
    if habit_row.freq_num == 1 and habit_row.freq_den == 1:
        return "every day"
    return f"{habit_row.freq_num}× / {habit_row.freq_den} days"


def _success_rate(habit_row, computed, frm: Date, to: Date, today: Date, completions: int) -> float | str:
    # Expected days clamp to the habit's first known entry; SKIP days excluded
    # (same semantics as csv_io._append_period_sheets applied to one bucket).
    known = get_known(computed)
    if not known:
        return ""
    start = max(frm, known[-1].date)
    end = min(to, today)
    active_days = 0
    day = start
    while day <= end:
        entry = computed.get(day)
        if entry is None or entry.value != SKIP:
            active_days += 1
        day += timedelta(days=1)
    expected = active_days * habit_row.freq_num / habit_row.freq_den
    if expected <= 0:
        return ""
    return min(1.0, completions / expected)


def _append_daily_sheet(workbook, label_headers, records, frm: Date, to: Date, success_fn) -> None:
    """Rows = (labels, habit_row, computed entries); columns = dates oldest-first + Total."""
    ws = workbook.active
    ws.title = "Daily"
    days = [frm + timedelta(days=i) for i in range((to - frm).days + 1)]
    ws.append([*label_headers, *[d.isoformat() for d in days], "Total"])
    day_totals = [0] * len(days)
    for labels, habit_row, computed in records:
        cells = []
        completions = 0
        total_value = 0.0
        for i, day in enumerate(days):
            entry = computed.get(day)
            value = entry.value if entry is not None else None
            cells.append(_day_cell(habit_row, value))
            if value is not None and success_fn(habit_row, value):
                completions += 1
                day_totals[i] += 1
            if habit_row.type == 1 and value is not None and value >= 0:
                total_value += value / 1000
        total = round(total_value, 2) if habit_row.type == 1 else completions
        ws.append([*labels, *cells, total])
    ws.append(["Totals", *[""] * (len(label_headers) - 1), *day_totals, sum(day_totals)])


def _save(workbook: Workbook) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


async def export_personal_xlsx(
    session: AsyncSession, user_id: int, from_date: Date | None, to_date: Date | None
) -> tuple[bytes, Date, Date]:
    prefs = await UserRepo(session).get_or_create_preferences(user_id)
    frm, to = _resolve_window(from_date, to_date, habit_math.user_today(prefs))
    rows = await HabitRepo(session).list_for_user(user_id, archived=False)
    entries = await EntryRepo(session).all_for_habits([r.id for r in rows])

    records = []
    for row in rows:
        habit = habit_math.to_domain(row)
        computed = habit_math.computed_entries_for(habit, entries[row.id])
        records.append(([row.name], row, computed))

    workbook = Workbook()
    _append_daily_sheet(workbook, ["Habit"], records, frm, to, _is_completed)
    _style_headers(workbook)
    workbook["Daily"].freeze_panes = "B2"
    return _save(workbook), frm, to


async def export_room_xlsx(
    session: AsyncSession, user_id: int, room_id: int, from_date: Date | None, to_date: Date | None
) -> tuple[bytes, Date, Date]:
    await _require_admin(session, room_id, user_id)
    user_repo = UserRepo(session)
    prefs = await user_repo.get_or_create_preferences(user_id)
    # One fixed calendar window for all members (from the caller's today); score and
    # streak are still computed at each member's own today, like the leaderboard.
    frm, to = _resolve_window(from_date, to_date, habit_math.user_today(prefs))

    repo = RoomRepo(session)
    members = await repo.members_with_users(room_id)
    room_habits = await repo.room_habits(room_id)
    links = await repo.links_for_room_habits([rh.id for rh in room_habits])
    habit_names = {rh.id: rh.name for rh in room_habits}
    links_by_user: dict[int, list] = {}
    for link in links:
        links_by_user.setdefault(link.user_id, []).append(link)

    habit_repo = HabitRepo(session)
    entry_repo = EntryRepo(session)

    records = []
    analytics = []
    for _membership, user in members:
        member_prefs = await user_repo.get_or_create_preferences(user.id)
        member_today = habit_math.user_today(member_prefs)
        label = f"{user.first_name} (@{user.username})" if user.username else user.first_name
        for link in links_by_user.get(user.id, []):
            try:
                habit_row = await habit_repo.get_owned(link.habit_id, user.id)
            except NotFoundError:
                continue
            habit = habit_math.to_domain(habit_row)
            entry_rows = await entry_repo.all_for_habit(habit_row.id)
            computed = habit_math.computed_entries_for(habit, entry_rows)
            name = habit_names.get(link.room_habit_id, habit_row.name)
            records.append(([label, name], habit_row, computed))
            completions = sum(
                1 for d, e in computed.items() if frm <= d <= to and _is_success(habit_row, e.value)
            )
            analytics.append([
                label,
                name,
                _format_target(habit_row),
                habit_math.score_on(habit, computed, member_today),
                habit_math.current_streak(habit, computed, member_today),
                completions,
                _success_rate(habit_row, computed, frm, to, member_today, completions),
            ])

    workbook = Workbook()
    _append_daily_sheet(workbook, ["Member", "Habit"], records, frm, to, _is_success)

    ws = workbook.create_sheet("Analytics")
    ws.append(["Member", "Habit", "Target", "Score", "Current streak", "Completions", "Success rate"])
    for row in analytics:
        ws.append(row)
        ws.cell(row=ws.max_row, column=4).number_format = "0.0%"
        ws.cell(row=ws.max_row, column=7).number_format = "0.0%"

    ws = workbook.create_sheet("Leaderboard")
    ws.append(["Rank", "Member", "Username", "Score", "Streak", "Completions", "Linked habits"])
    for rank, entry in enumerate(await leaderboard(session, user_id, room_id, window=(frm, to)), 1):
        ws.append([
            rank, entry.first_name, entry.username or "",
            entry.score, entry.streak, entry.completions, entry.linked_habits,
        ])
        ws.cell(row=ws.max_row, column=4).number_format = "0%"

    _style_headers(workbook)
    workbook["Daily"].freeze_panes = "C2"
    return _save(workbook), frm, to
