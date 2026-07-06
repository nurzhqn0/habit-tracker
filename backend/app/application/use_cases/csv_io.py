"""Loop-format CSV export/import and Excel export.

Export ZIP layout mirrors uhabits HabitsCSVExporter: Habits.csv at the root,
plus "NNN <name>/Checkmarks.csv" and "NNN <name>/Scores.csv" per habit.
Import consumes the same layout (matches by habit name; entries upserted).
"""

import csv
import io
import re
import zipfile
from datetime import date as Date

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.ext.asyncio import AsyncSession

from app.application import habit_math
from app.domain.errors import ValidationError
from app.domain.models.entry import NO, SKIP, UNKNOWN, YES_AUTO, YES_MANUAL
from app.domain.services.entry_list import get_known
from app.infrastructure.repositories.habit_repo import EntryRepo, HabitRepo
from app.infrastructure.repositories.user_repo import UserRepo

PALETTE_HEX = [
    "#D32F2F", "#E64A19", "#F57C00", "#FF8F00", "#F9A825",
    "#AFB42B", "#7CB342", "#388E3C", "#00897B", "#00ACC1",
    "#039BE5", "#1976D2", "#303F9F", "#5E35B1", "#8E24AA",
    "#D81B60", "#5D4037", "#424242", "#757575", "#9E9E9E",
]

VALUE_NAMES = {YES_MANUAL: "YES_MANUAL", YES_AUTO: "YES_AUTO", NO: "NO", SKIP: "SKIP", UNKNOWN: "UNKNOWN"}
NAMES_TO_VALUE = {v: k for k, v in VALUE_NAMES.items()}

MAX_UPLOAD_BYTES = 5 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 30 * 1024 * 1024
MAX_HABITS = 200
MAX_ENTRIES_PER_HABIT = 20_000

HABITS_HEADER = [
    "Position", "Name", "Type", "Question", "Description", "FrequencyNumerator",
    "FrequencyDenominator", "Color", "Unit", "Target Type", "Target Value", "Archived?",
]


def _csv(rows: list[list[str]]) -> str:
    buffer = io.StringIO()
    csv.writer(buffer, lineterminator="\n").writerows(rows)
    return buffer.getvalue()


def _sanitize(name: str) -> str:
    return re.sub(r"[^ a-zA-Z0-9._-]+", "", name)[:100].strip()


def _format_value(habit_type: int, value: int) -> str:
    if habit_type == 0 and value in VALUE_NAMES:
        return VALUE_NAMES[value]
    return str(value)


async def _load(session: AsyncSession, user_id: int, habit_id: int | None):
    repo = HabitRepo(session)
    if habit_id is not None:
        rows = [await repo.get_owned(habit_id, user_id)]
    else:
        rows = await repo.list_for_user(user_id)
    entries = await EntryRepo(session).all_for_habits([r.id for r in rows])
    prefs = await UserRepo(session).get_or_create_preferences(user_id)
    return rows, entries, habit_math.user_today(prefs)


def _habits_csv(rows) -> str:
    table = [HABITS_HEADER]
    for index, row in enumerate(rows):
        table.append([
            f"{index + 1:03d}",
            row.name,
            "YES_NO" if row.type == 0 else "NUMERICAL",
            row.question,
            row.description,
            str(row.freq_num),
            str(row.freq_den),
            PALETTE_HEX[row.color] if 0 <= row.color < 20 else PALETTE_HEX[8],
            row.unit if row.type == 1 else "",
            ("AT_LEAST" if row.target_type == 0 else "AT_MOST") if row.type == 1 else "",
            f"{row.target_value:.1f}" if row.type == 1 else "",
            "1" if row.archived else "0",
        ])
    return _csv(table)


async def export_zip(session: AsyncSession, user_id: int, habit_id: int | None = None) -> bytes:
    rows, entries_by_habit, today = await _load(session, user_id, habit_id)

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("Habits.csv", _habits_csv(rows))
        for index, row in enumerate(rows):
            habit = habit_math.to_domain(row)
            computed = habit_math.computed_entries_for(habit, entries_by_habit[row.id])
            dir_name = f"{index + 1:03d} {_sanitize(row.name)}"

            checkmarks = [["Date", "Value", "Notes"]]
            for entry in get_known(computed):
                checkmarks.append([entry.date.isoformat(), _format_value(row.type, entry.value), entry.notes])
            zf.writestr(f"{dir_name}/Checkmarks.csv", _csv(checkmarks))

            scores = [["Date", "Score"]]
            for score in reversed(habit_math.scores_for(habit, computed, today)):
                scores.append([score.date.isoformat(), f"{score.value:.4f}"])
            zf.writestr(f"{dir_name}/Scores.csv", _csv(scores))
    return buffer.getvalue()


def _is_completed(row, value: int) -> bool:
    if row.type == 0:
        return value in (YES_AUTO, YES_MANUAL)
    if value < 0:
        return False
    actual = value / 1000
    return actual <= row.target_value if row.target_type == 1 else actual >= row.target_value


def _week_key(date: Date) -> str:
    iso = date.isocalendar()
    return f"{iso[0]}-W{iso[1]:02d}"


def _month_key(date: Date) -> str:
    return f"{date.year}-{date.month:02d}"


def _append_summary(workbook: Workbook, prepared: list, today: Date) -> None:
    ws = workbook.create_sheet("Summary")
    ws.append([
        "Habit", "Type", "Success rate", "Current streak", "Best streak",
        "Total completions", "Total value", "Daily average", "First entry", "Last entry",
    ])
    for row, habit, computed, scores in prepared:
        known = get_known(computed)  # newest first
        completions = sum(1 for e in known if _is_completed(row, e.value))
        entry_days = sum(1 for e in known if e.value >= 0)
        total_value = sum(e.value for e in known if e.value >= 0) / 1000 if row.type == 1 else None
        streaks = habit_math.streaks_for(habit, computed, today)
        ws.append([
            row.name,
            "YES_NO" if row.type == 0 else "NUMERICAL",
            habit_math.score_on(habit, computed, today),
            habit_math.current_streak(habit, computed, today),
            max((s.length for s in streaks), default=0),
            completions,
            round(total_value, 2) if total_value is not None else "",
            round(total_value / entry_days, 2) if total_value is not None and entry_days else "",
            known[-1].date.isoformat() if known else "",
            known[0].date.isoformat() if known else "",
        ])
        ws.cell(row=ws.max_row, column=3).number_format = "0.0%"


def _append_period_sheets(workbook: Workbook, prepared: list) -> None:
    for sheet_name, period_header, key_fn in (
        ("Weekly", "Week", _week_key),
        ("Monthly", "Month", _month_key),
    ):
        ws = workbook.create_sheet(sheet_name)
        ws.append(["Habit", period_header, "Completions", "Total value", "Avg score"])
        for row, habit, computed, scores in prepared:
            buckets: dict[str, dict] = {}
            for entry in get_known(computed):
                bucket = buckets.setdefault(key_fn(entry.date), {"completions": 0, "total": 0.0})
                if _is_completed(row, entry.value):
                    bucket["completions"] += 1
                if row.type == 1 and entry.value >= 0:
                    bucket["total"] += entry.value / 1000
            score_buckets: dict[str, list[float]] = {}
            for date, value in scores.items():
                score_buckets.setdefault(key_fn(date), []).append(value)
            for key in sorted(buckets):
                values = score_buckets.get(key, [])
                ws.append([
                    row.name,
                    key,
                    buckets[key]["completions"],
                    round(buckets[key]["total"], 2) if row.type == 1 else "",
                    sum(values) / len(values) if values else 0.0,
                ])
                ws.cell(row=ws.max_row, column=5).number_format = "0.0%"


def _style_headers(workbook: Workbook) -> None:
    for ws in workbook.worksheets:
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"


async def export_xlsx(session: AsyncSession, user_id: int, habit_id: int | None = None) -> bytes:
    rows, entries_by_habit, today = await _load(session, user_id, habit_id)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Habits"
    sheet.append(HABITS_HEADER)
    for index, row in enumerate(rows):
        sheet.append([
            index + 1, row.name, "YES_NO" if row.type == 0 else "NUMERICAL", row.question,
            row.description, row.freq_num, row.freq_den,
            PALETTE_HEX[row.color] if 0 <= row.color < 20 else PALETTE_HEX[8],
            row.unit if row.type == 1 else "",
            ("AT_LEAST" if row.target_type == 0 else "AT_MOST") if row.type == 1 else "",
            row.target_value if row.type == 1 else "",
            "yes" if row.archived else "no",
        ])

    prepared = []
    for row in rows:
        habit = habit_math.to_domain(row)
        computed = habit_math.computed_entries_for(habit, entries_by_habit[row.id])
        scores = {s.date: s.value for s in habit_math.scores_for(habit, computed, today)}
        prepared.append((row, habit, computed, scores))

    _append_summary(workbook, prepared, today)
    _append_period_sheets(workbook, prepared)

    used_titles = {"Habits", "Summary", "Weekly", "Monthly"}
    for index, (row, habit, computed, scores) in enumerate(prepared):
        title = _sanitize(row.name)[:24] or f"Habit {index + 1}"
        base = title
        suffix = 2
        while title in used_titles:
            title = f"{base[:20]} ({suffix})"
            suffix += 1
        used_titles.add(title)

        ws = workbook.create_sheet(title)
        ws.append(["Date", "Value", "Real value", "Notes", "Score"])
        for entry in get_known(computed):
            ws.append([
                entry.date.isoformat(),
                _format_value(row.type, entry.value),
                entry.value / 1000 if row.type == 1 and entry.value >= 0 else None,
                entry.notes,
                round(scores.get(entry.date, 0.0), 4),
            ])

    _style_headers(workbook)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _parse_value(raw: str, habit_type: int) -> int | None:
    raw = raw.strip()
    if raw in NAMES_TO_VALUE:
        return NAMES_TO_VALUE[raw]
    try:
        return int(float(raw))
    except ValueError:
        return None


async def import_zip(session: AsyncSession, user_id: int, data: bytes) -> dict:
    if len(data) > MAX_UPLOAD_BYTES:
        raise ValidationError("Archive too large")

    try:
        zf = zipfile.ZipFile(io.BytesIO(data))
    except zipfile.BadZipFile as e:
        raise ValidationError("Not a valid ZIP archive") from e

    if sum(info.file_size for info in zf.infolist()) > MAX_UNCOMPRESSED_BYTES:
        raise ValidationError("Archive contents too large")

    names = zf.namelist()
    habits_csv = next((n for n in names if n.endswith("Habits.csv") and "/" not in n.rstrip("Habits.csv")), None)
    if habits_csv is None:
        habits_csv = next((n for n in names if n.endswith("Habits.csv")), None)
    if habits_csv is None:
        raise ValidationError("Habits.csv not found in archive")

    reader = csv.reader(io.StringIO(zf.read(habits_csv).decode("utf-8-sig")))
    rows = list(reader)
    if not rows or [c.strip() for c in rows[0][:2]] != ["Position", "Name"]:
        raise ValidationError("Habits.csv has an unexpected format")
    if len(rows) - 1 > MAX_HABITS:
        raise ValidationError("Too many habits in archive")

    habit_repo = HabitRepo(session)
    entry_repo = EntryRepo(session)
    existing = {h.name: h for h in await habit_repo.list_for_user(user_id)}
    hex_to_color = {hex_value.lower(): i for i, hex_value in enumerate(PALETTE_HEX)}

    created = 0
    updated_entries = 0

    for row in rows[1:]:
        if len(row) < 12 or not row[1].strip():
            continue
        position, name, type_name = row[0], row[1].strip(), row[2].strip()
        habit_type = 1 if type_name == "NUMERICAL" else 0

        habit = existing.get(name)
        if habit is None:
            habit = await habit_repo.create(
                user_id,
                name=name,
                question=row[3],
                description=row[4],
                type=habit_type,
                freq_num=max(1, int(float(row[5] or 1))),
                freq_den=max(1, int(float(row[6] or 1))),
                color=hex_to_color.get(row[7].strip().lower(), 8),
                unit=row[8],
                target_type=1 if row[9].strip() == "AT_MOST" else 0,
                target_value=float(row[10] or 0),
                archived=row[11].strip() in ("1", "yes", "true"),
            )
            existing[name] = habit
            created += 1

        prefix = f"{int(position):03d} " if position.strip().isdigit() else ""
        dir_prefix = f"{prefix}{_sanitize(name)}"
        checkmarks_name = next(
            (n for n in names if n.endswith("/Checkmarks.csv") and dir_prefix in n), None
        )
        if checkmarks_name is None:
            continue

        entry_reader = csv.reader(io.StringIO(zf.read(checkmarks_name).decode("utf-8-sig")))
        entry_rows = list(entry_reader)[1:]
        if len(entry_rows) > MAX_ENTRIES_PER_HABIT:
            raise ValidationError(f"Too many entries for habit {name}")

        for entry_row in entry_rows:
            if len(entry_row) < 2:
                continue
            try:
                entry_date = Date.fromisoformat(entry_row[0].strip())
            except ValueError:
                continue
            value = _parse_value(entry_row[1], habit_type)
            notes = entry_row[2] if len(entry_row) > 2 else ""
            # YES_AUTO and UNKNOWN are derived — never persist them.
            if value is None or value in (YES_AUTO, UNKNOWN):
                continue
            await entry_repo.upsert(habit.id, entry_date, value, notes or None)
            updated_entries += 1

    await session.commit()
    return {"habits_created": created, "entries_imported": updated_entries}
